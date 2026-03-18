"""
שליחת אימיילים המוניים עם קובץ מצורף
תומך ב: רשימת אימיילים מ-Excel, HTML template, קבצים מצורפים
"""

import smtplib
import time
import os
import sys
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import openpyxl
import config


def load_emails(filepath: str, column_name: str) -> list[str]:
    """קורא רשימת אימיילים מקובץ Excel"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    if column_name not in headers:
        raise ValueError(f"עמודה '{column_name}' לא נמצאה. עמודות קיימות: {headers}")

    col_index = headers.index(column_name)
    emails = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        email = row[col_index]
        if email and "@" in str(email):
            emails.append(str(email).strip())

    return emails


def load_template(filepath: str, name: str = None) -> str:
    """טוען את תבנית ה-HTML (עם תמיכה עתידית בפרסונליזציה)"""
    with open(filepath, "r", encoding="utf-8") as f:
        html = f.read()
    # בעתיד: html = html.replace("{{name}}", name or "")
    return html


def build_email(sender: str, recipient: str, subject: str, html_body: str, attachments: list) -> MIMEMultipart:
    """בונה את אובייקט האימייל"""
    msg = MIMEMultipart("mixed")
    msg["From"] = sender
    msg["To"] = recipient
    msg["Subject"] = subject

    # תוכן HTML
    msg.attach(MIMEText(html_body, "html", "utf-8"))

    # קבצים מצורפים
    for filepath in attachments:
        if not os.path.exists(filepath):
            print(f"  אזהרה: קובץ מצורף לא נמצא: {filepath}")
            continue
        with open(filepath, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        filename = os.path.basename(filepath)
        part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
        msg.attach(part)

    return msg


def send_all():
    # טעינת נתונים
    print("טוען רשימת אימיילים...")
    emails = load_emails(config.CONTACTS_FILE, config.EMAIL_COLUMN)
    print(f"נמצאו {len(emails)} כתובות אימייל")

    html_body = load_template("template.html")

    # התחברות ל-Gmail
    print("\nמתחבר ל-Gmail...")
    try:
        server = smtplib.SMTP_SSL("smtp.gmail.com", 465)
        server.login(config.SENDER_EMAIL, config.SENDER_PASSWORD)
        print("התחברות הצליחה!\n")
    except Exception as e:
        print(f"שגיאת התחברות: {e}")
        print("ודא שה-App Password נכון ו-2FA מופעל בחשבון Google")
        sys.exit(1)

    # שליחה
    success, failed = 0, []

    for i, email in enumerate(emails, 1):
        try:
            msg = build_email(
                sender=config.SENDER_EMAIL,
                recipient=email,
                subject=config.SUBJECT,
                html_body=html_body,
                attachments=config.ATTACHMENTS
            )
            server.sendmail(config.SENDER_EMAIL, email, msg.as_string())
            print(f"[{i}/{len(emails)}] נשלח ל: {email}")
            success += 1
            time.sleep(config.DELAY_BETWEEN_EMAILS)

        except Exception as e:
            print(f"[{i}/{len(emails)}] נכשל: {email} — {e}")
            failed.append(email)

    server.quit()

    # סיכום
    print(f"\n{'='*40}")
    print(f"נשלח בהצלחה: {success}")
    print(f"נכשל:        {len(failed)}")
    if failed:
        print("\nכתובות שנכשלו:")
        for f in failed:
            print(f"  - {f}")


if __name__ == "__main__":
    send_all()
